"""Parser for the official Swiss Parliament session-vote XLSX layouts."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

import openpyxl


HEADER_MARKERS = {"VoteDate", "Abstimmungsdatum", "Geschäftsnummer"}
AGGREGATE_MARKERS = {"Decision", "Entscheid des Rates"}
MEMBER_LABELS = {
    "CouncillorId",
    "CouncillorBioId",
    "CouncillorName",
    "Ratsmitglied (Nr)",
    "Name des Ratsmitgliedes",
    "Name des Ratsmitglieds",
}


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "*":
        return None
    return text


def normalized_name(value: Any) -> str:
    text = clean_text(value) or ""
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        text = f"{first} {last}"
    text = unicodedata.normalize("NFKD", text.casefold())
    text = "".join(character for character in text if not unicodedata.combining(character))
    return " ".join(re.sub(r"[^a-z0-9]+", " ", text).split())


def display_name(value: Any) -> str:
    text = clean_text(value) or "Unbekannte Person"
    if "," not in text:
        return text
    last, first = [part.strip() for part in text.split(",", 1)]
    return f"{first} {last}".strip()


def canonical_affair_identifier(value: Any, vote_date: str) -> tuple[str | None, str | None]:
    text = clean_text(value)
    if text is None:
        return None, None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    if text.isdigit() and len(text) == 8:
        return text, f"{text[:2]}.{int(text[4:]):04d}"
    match = re.fullmatch(r"(\d{2})\.(\d{1,4})", text)
    if match:
        short_year = int(match.group(1))
        vote_year = int(vote_date[:4])
        century = vote_year // 100
        full_year = century * 100 + short_year
        if full_year > vote_year + 1:
            full_year -= 100
        serial = int(match.group(2))
        return f"{full_year:04d}{serial:04d}", f"{short_year:02d}.{serial:04d}"
    return text, text


def iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if text is None:
        raise ValueError("Voting row has no date")
    for pattern in ("%d.%m.%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.isoformat(timespec="seconds") if "H" in pattern else parsed.date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Unsupported voting date {text!r}")


def normalized_choice(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    token = normalized_name(text)
    if token in {"ja", "yes"}:
        return "yes"
    if token in {"nein", "no"}:
        return "no"
    if token in {"eh", "enth", "enthaltung"} or "enthaltung" in token:
        return "abstain"
    if token in {"es"} or "entschuld" in token:
        return "excused"
    if token in {"nt"} or "nicht teilgenommen" in token:
        return "not_participating"
    if token in {"p"} or "prasident" in token:
        return "presiding"
    return "other"


def normalized_aggregate_label(value: Any) -> str | None:
    """Map descriptive workbook aggregate headers to canonical choices.

    Current workbooks prefix the decision with ``Anzahl`` (for example,
    ``Anzahl 'Ja'``), while older layouts use the bare decision label.  These
    are column labels rather than individual member decisions, so unknown
    headers must be ignored instead of becoming an ``other`` aggregate.
    """

    token = normalized_name(value)
    if not token:
        return None
    if token.startswith("anzahl "):
        token = token.removeprefix("anzahl ").strip()
    if token in {"ja", "yes"}:
        return "yes"
    if token in {"nein", "no"}:
        return "no"
    if "enthaltung" in token:
        return "abstain"
    if "entschuld" in token:
        return "excused"
    if "nicht teilgenommen" in token:
        return "not_participating"
    if "prasident" in token:
        return "presiding"
    return None


def normalized_vote_type(value: Any) -> tuple[str, str]:
    token = normalized_name(value)
    if any(phrase in token for phrase in ("schlussabstimmung", "vote final")):
        return "final", "derived_from_official_question_text"
    if any(phrase in token for phrase in ("gesamtabstimmung", "vote sur l ensemble")):
        return "overall", "derived_from_official_question_text"
    if any(phrase in token for phrase in ("eintreten", "entrer en matiere")):
        return "entry", "derived_from_official_question_text"
    if "dringlich" in token or "urgence" in token:
        return "urgency", "derived_from_official_question_text"
    return "other", "unclassified_official_question_text"


def _find_layout(rows: list[tuple[Any, ...]]) -> tuple[str, int, tuple[Any, ...]]:
    for index, row in enumerate(rows, start=1):
        values = {str(value).strip() for value in row if value is not None}
        if "VoteDate" in values:
            return "legacy", index, row
        if "Abstimmungsdatum" in values:
            return "transitional", index, row
        if "Geschäftsnummer" in values:
            return "current", index, row
    raise ValueError("No supported vote header row found in first 15 workbook rows")


def _header_index(header: tuple[Any, ...], *names: str) -> int | None:
    for index, value in enumerate(header):
        if clean_text(value) in names:
            return index
    return None


def _metadata(rows: list[tuple[Any, ...]], header_row: int) -> tuple[int, dict[int, dict[str, Any]]]:
    labels: dict[str, tuple[int, tuple[Any, ...]]] = {}
    member_start = None
    for row in rows[: header_row - 1]:
        for column, value in enumerate(row):
            label = clean_text(value)
            if label in MEMBER_LABELS or label in {
                "Rat",
                "Fraktion",
                "Kanton",
                "Geschlecht",
                "Geburtsdatum",
                "Vereidigungsdatum",
            }:
                labels[label] = (column, row)
                if label in MEMBER_LABELS:
                    member_start = column + 1
    if member_start is None:
        raise ValueError("Workbook contains no member metadata header")

    members: dict[int, dict[str, Any]] = {}
    label_map = {
        "CouncillorId": "voting_councillor_number",
        "CouncillorBioId": "cv_person_id",
        "CouncillorName": "name",
        "Ratsmitglied (Nr)": "cv_person_id",
        "Name des Ratsmitgliedes": "name",
        "Name des Ratsmitglieds": "name",
        "Rat": "chamber",
        "Fraktion": "faction",
        "Kanton": "canton",
        "Geschlecht": "gender",
        "Geburtsdatum": "birth_date",
        "Vereidigungsdatum": "oath_date",
    }
    for label, (_label_column, row) in labels.items():
        field = label_map[label]
        for column in range(member_start, len(row)):
            value = row[column]
            if clean_text(value) is not None:
                members.setdefault(column, {})[field] = value
    for member in members.values():
        member["display_name"] = display_name(member.get("name"))
        member["normalized_name"] = normalized_name(member.get("name"))
        for field in ("birth_date", "oath_date"):
            value = member.get(field)
            if isinstance(value, datetime):
                member[field] = value.date().isoformat()
            elif isinstance(value, date):
                member[field] = value.isoformat()
            elif value is not None:
                member[field] = clean_text(value)
    return member_start, members


def parse_workbook(path: Path, chamber_hint: str | None = None) -> Iterator[dict[str, Any]]:
    """Yield one canonical source record for each official workbook vote row."""

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        preview = list(
            worksheet.iter_rows(min_row=1, max_row=15, max_col=300, values_only=True)
        )
        layout, header_row, header = _find_layout(preview)
        member_start, members = _metadata(preview, header_row)
        aggregate_start = _header_index(header, *AGGREGATE_MARKERS)
        if aggregate_start is None:
            raise ValueError("Workbook contains no aggregate-result columns")
        members = {
            column: member for column, member in members.items() if member_start <= column < aggregate_start
        }

        if layout == "current":
            fields = {
                "affair": _header_index(header, "Geschäftsnummer"),
                "title": _header_index(header, "Geschäftstitel"),
                "affair_type": _header_index(header, "Geschäftstyp"),
                "draft_number": _header_index(header, "Entwurfnummer"),
                "submission": _header_index(header, "Entwurftitel"),
                "registration": _header_index(header, "Referenznummer"),
                "date": _header_index(header, "Datum der Abstimmung"),
                "division": _header_index(header, "Abstimmungsgegenstand"),
                "yes": _header_index(header, "Bedeutung Ja"),
                "no": _header_index(header, "Bedeutung Nein"),
                "chamber": None,
                "committee": None,
                "department": None,
            }
        else:
            fields = {
                "affair": _header_index(header, "AffairId", "Geschäftsnummer"),
                "title": _header_index(header, "AffairTitle", "Geschäftstitel"),
                "affair_type": None,
                "draft_number": None,
                "submission": _header_index(header, "VoteSubmissionText", "Vorlagetitel"),
                "registration": _header_index(header, "VoteRegistrationNumber", "Referenznummer"),
                "date": _header_index(header, "VoteDate", "Abstimmungsdatum"),
                "division": _header_index(header, "DivisionText", "Abstimmungsgegenstand"),
                "yes": _header_index(header, "VoteMeaningYes", "Bedeutung Ja"),
                "no": _header_index(header, "VoteMeaningNo", "Bedeutung Nein"),
                "chamber": _header_index(header, "Rat"),
                "committee": _header_index(header, "Kommission", "Zuständige Kommission"),
                "department": _header_index(header, "Dept.", "Zuständige Behörde"),
            }
        if fields["registration"] is None or fields["date"] is None:
            raise ValueError("Workbook lacks a vote registration or date column")

        aggregate_headers = {
            column: clean_text(value) for column, value in enumerate(header) if column >= aggregate_start
        }
        declared_count = None
        for row in preview[: header_row - 1]:
            for column, value in enumerate(row[:-1]):
                if clean_text(value) == "Anzahl Abstimmungen:":
                    try:
                        declared_count = int(row[column + 1])
                    except (TypeError, ValueError):
                        pass

        source_rows = list(worksheet.iter_rows(min_row=header_row + 1, values_only=True))
        repaired_rows: list[tuple[int, tuple[Any, ...]]] = []
        for row_number, row in enumerate(source_rows, start=header_row + 1):
            date_value = row[fields["date"]] if fields["date"] < len(row) else None
            date_is_valid = True
            try:
                iso_date(date_value)
            except ValueError:
                date_is_valid = False
            if layout == "legacy" and clean_text(date_value) and not date_is_valid:
                # One official 2015 workbook splits a long voting row across two
                # spreadsheet rows. The continuation starts with parenthesized
                # question text, then a submission title, then choices shifted
                # nine columns left. Repair that exact structural defect while
                # retaining both physical row numbers in source provenance.
                first = clean_text(row[fields["date"]]) or ""
                if not first.startswith("(") or not repaired_rows:
                    raise ValueError(
                        f"Unexpected non-date continuation at {path.name}:{row_number}"
                    )
                previous_number, previous = repaired_rows.pop()
                merged = list(previous)
                while len(merged) < len(row) + (member_start - 3):
                    merged.append(None)
                division_column = fields["division"]
                submission_column = fields["submission"]
                merged[division_column] = " ".join(
                    part for part in (clean_text(merged[division_column]), clean_text(row[0])) if part
                )
                merged[submission_column] = clean_text(row[1]) or merged[submission_column]
                shift = member_start - 3
                for source_column in range(3, len(row)):
                    if row[source_column] is not None:
                        merged[source_column + shift] = row[source_column]
                repaired_rows.append((previous_number, tuple(merged)))
                continue
            repaired_rows.append((row_number, row))

        actual_count = 0
        for row_number, row in repaired_rows:
            if not row or all(clean_text(value) is None for value in row[:aggregate_start]):
                continue
            registration = clean_text(row[fields["registration"]])
            if registration is None:
                continue
            occurred_at = iso_date(row[fields["date"]])
            raw_affair = row[fields["affair"]] if fields["affair"] is not None else None
            affair_id, formatted_affair_id = canonical_affair_identifier(raw_affair, occurred_at)
            chamber = chamber_hint
            if fields["chamber"] is not None:
                chamber = clean_text(row[fields["chamber"]]) or chamber
            if chamber not in {"NR", "SR"}:
                raise ValueError(f"Unsupported or missing chamber {chamber!r} in {path.name}")

            row_shift = 0
            if layout == "legacy":
                for candidate_shift in range(0, 12):
                    start = aggregate_start + candidate_shift
                    if start + 5 >= len(row):
                        continue
                    decision_token = normalized_choice(row[start])
                    counts_are_numeric = all(
                        isinstance(row[start + offset], (int, float))
                        for offset in range(1, 6)
                    )
                    if decision_token in {"yes", "no"} and counts_are_numeric:
                        row_shift = candidate_shift
                        break

            choices = []
            for column, metadata in members.items():
                shifted_column = column + row_shift
                decision = row[shifted_column] if shifted_column < len(row) else None
                choice = normalized_choice(decision)
                if choice is None:
                    # An empty member cell carries no explicit individual
                    # decision. Do not infer one from the aggregate because
                    # some workbooks also include not-yet-active members as
                    # empty columns, making the missing individual ambiguous.
                    continue
                choices.append(
                    {
                        **metadata,
                        "raw_decision": clean_text(decision),
                        "normalized_choice": choice,
                        "column": column + 1,
                    }
                )

            aggregates: dict[str, int] = {}
            overall_decision = None
            for column, label in aggregate_headers.items():
                shifted_column = column + row_shift
                value = row[shifted_column] if shifted_column < len(row) else None
                if label in AGGREGATE_MARKERS:
                    overall_decision = clean_text(value)
                    continue
                normalized = normalized_aggregate_label(label)
                if normalized and value not in (None, ""):
                    try:
                        aggregates[normalized] = int(value)
                    except (TypeError, ValueError):
                        continue

            division = row[fields["division"]] if fields["division"] is not None else None
            submission = row[fields["submission"]] if fields["submission"] is not None else None
            if layout == "legacy" and row_shift:
                submission_column = member_start + row_shift - 2
                submission = row[submission_column]
                division = " ".join(
                    str(value).strip()
                    for value in row[fields["division"] : submission_column]
                    if clean_text(value)
                )
            vote_type, vote_type_basis = normalized_vote_type(division)
            actual_count += 1
            yield {
                "source_format": "official_session_xlsx",
                "workbook": path.name,
                "worksheet": worksheet.title,
                "row_number": row_number,
                "layout": layout,
                "declared_workbook_vote_count": declared_count,
                "chamber": chamber,
                "affair_id": affair_id,
                "affair_formatted_id": formatted_affair_id,
                "affair_title": clean_text(row[fields["title"]]) if fields["title"] is not None else None,
                "affair_type": clean_text(row[fields["affair_type"]]) if fields["affair_type"] is not None else None,
                "draft_number": clean_text(row[fields["draft_number"]]) if fields["draft_number"] is not None else None,
                "registration_number": registration,
                "occurred_at": occurred_at,
                "committee": clean_text(row[fields["committee"]]) if fields["committee"] is not None else None,
                "department": clean_text(row[fields["department"]]) if fields["department"] is not None else None,
                "division_text": clean_text(division),
                "submission_text": clean_text(submission),
                "meaning_yes": clean_text(row[fields["yes"]]) if fields["yes"] is not None else None,
                "meaning_no": clean_text(row[fields["no"]]) if fields["no"] is not None else None,
                "overall_decision": overall_decision,
                "vote_type": vote_type,
                "vote_type_basis": vote_type_basis,
                "aggregates": aggregates,
                "choices": choices,
            }
        if declared_count is not None and actual_count != declared_count:
            raise ValueError(
                f"Workbook {path.name} declares {declared_count} votes but contains {actual_count} parsed rows"
            )
    finally:
        workbook.close()
